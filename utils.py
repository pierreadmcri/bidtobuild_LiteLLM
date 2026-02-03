"""
Fonctions utilitaires pour la gestion des erreurs, retry, validation et rate limiting
"""
import os
import time
import logging
from pathlib import Path
from functools import wraps
from typing import Callable, Any, List, Union
from openai import OpenAI
from PIL import Image
import pytesseract
import config

# Configuration du logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==========================================
# VALIDATION & SÉCURITÉ
# ==========================================

class ValidationError(Exception):
    """Exception levée lors d'une erreur de validation"""
    pass

class FileTooLargeError(Exception):
    """Exception levée quand un fichier dépasse la taille maximale"""
    pass

def validate_file_path(path: str) -> Path:
    """
    Valide un chemin de fichier/dossier pour éviter les path traversal attacks

    Args:
        path: Chemin à valider

    Returns:
        Path object résolu et validé

    Raises:
        ValidationError: Si le chemin est invalide ou dangereux
    """
    try:
        # Résoudre le chemin absolu
        resolved_path = Path(path).resolve()

        # Vérifier que le chemin existe
        if not resolved_path.exists():
            raise ValidationError(f"Le chemin n'existe pas : {path}")

        # Vérifier les patterns interdits
        path_str = str(resolved_path)
        for forbidden in config.FORBIDDEN_PATH_PATTERNS:
            if forbidden.lower() in path_str.lower():
                raise ValidationError(f"Accès interdit à ce chemin : {path}")

        logger.info(f"Chemin validé : {resolved_path}")
        return resolved_path

    except Exception as e:
        if isinstance(e, ValidationError):
            raise
        raise ValidationError(f"Erreur de validation du chemin : {str(e)}")

def validate_file_size(filepath: Path) -> None:
    """
    Vérifie que la taille du fichier ne dépasse pas la limite

    Args:
        filepath: Chemin du fichier à vérifier

    Raises:
        FileTooLargeError: Si le fichier est trop volumineux
    """
    try:
        file_size = filepath.stat().st_size
        max_size = config.MAX_FILE_SIZE_BYTES

        if file_size > max_size:
            size_mb = file_size / (1024 * 1024)
            max_mb = max_size / (1024 * 1024)
            raise FileTooLargeError(
                f"Fichier trop volumineux : {size_mb:.2f} MB (max: {max_mb:.2f} MB)"
            )

    except FileTooLargeError:
        raise
    except Exception as e:
        logger.warning(f"Impossible de vérifier la taille du fichier : {e}")

# ==========================================
# RETRY LOGIC avec EXPONENTIAL BACKOFF
# ==========================================

def retry_with_exponential_backoff(
    max_retries: int = config.MAX_RETRIES,
    base_delay: float = config.RETRY_BASE_DELAY,
    max_delay: float = config.RETRY_MAX_DELAY,
    exceptions: tuple = (Exception,)
):
    """
    Décorateur pour implémenter retry avec exponential backoff

    Args:
        max_retries: Nombre maximum de tentatives
        base_delay: Délai initial en secondes
        max_delay: Délai maximum en secondes
        exceptions: Tuple d'exceptions à intercepter

    Returns:
        Fonction décorée avec retry logic
    """
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs) -> Any:
            attempt = 0
            last_exception = None

            while attempt < max_retries:
                try:
                    return func(*args, **kwargs)

                except exceptions as e:
                    attempt += 1
                    last_exception = e

                    if attempt >= max_retries:
                        logger.error(
                            f"{func.__name__} a échoué après {max_retries} tentatives: {e}"
                        )
                        raise

                    # Calcul du délai avec exponential backoff
                    delay = min(base_delay * (2 ** (attempt - 1)), max_delay)

                    logger.warning(
                        f"{func.__name__} tentative {attempt}/{max_retries} échouée. "
                        f"Nouvelle tentative dans {delay}s: {e}"
                    )

                    time.sleep(delay)

            raise last_exception

        return wrapper
    return decorator

# ==========================================
# RATE LIMITER
# ==========================================

class RateLimiter:
    """
    Rate limiter simple pour éviter de dépasser les quotas API
    """
    def __init__(self, delay: float = config.RATE_LIMIT_DELAY):
        self.delay = delay
        self.last_call = 0

    def wait(self):
        """Attend le délai nécessaire avant le prochain appel"""
        current_time = time.time()
        elapsed = current_time - self.last_call

        if elapsed < self.delay:
            wait_time = self.delay - elapsed
            time.sleep(wait_time)

        self.last_call = time.time()

# Instance globale du rate limiter
rate_limiter = RateLimiter()

# ==========================================
# CLIENT OPENAI
# ==========================================

# Créer le client OpenAI avec la configuration du proxy
client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY"),
    base_url=config.OPENAI_API_BASE
)

# ==========================================
# WRAPPERS API avec RETRY & RATE LIMITING
# ==========================================

@retry_with_exponential_backoff()
def safe_completion(model: str = None, messages: List[dict] = None, **kwargs):
    """
    Wrapper sécurisé pour les appels de completion avec retry et rate limiting

    Utilise le client OpenAI natif avec la syntaxe client.chat.completions.create()

    Args:
        model: Nom du modèle (ex: "openai/gpt-4.1-mini")
        messages: Liste des messages pour le chat
        **kwargs: Arguments supplémentaires (temperature, stream, etc.)

    Returns:
        Réponse de l'API

    Raises:
        Exception: Si tous les retries échouent
    """
    rate_limiter.wait()

    model = model or config.MODEL_NAME

    try:
        # Utilisation de la syntaxe client OpenAI native
        response = client.chat.completions.create(
            model=model,
            messages=messages,
            **kwargs
        )
        logger.info(f"Completion réussie avec le modèle {model}")
        return response

    except Exception as e:
        logger.error(f"Erreur lors de l'appel completion: {e}")
        raise

@retry_with_exponential_backoff()
def safe_embedding(texts: List[str], model: str = None, **kwargs):
    """
    Wrapper sécurisé pour les embeddings avec retry et rate limiting

    Utilise le client OpenAI natif avec la syntaxe client.embeddings.create()

    Args:
        texts: Liste de textes à embedder (ou texte unique)
        model: Nom du modèle d'embedding (ex: "openai/text-embedding-3-small")
        **kwargs: Arguments supplémentaires

    Returns:
        Liste d'embeddings

    Raises:
        Exception: Si tous les retries échouent
    """
    rate_limiter.wait()

    model = model or config.EMBEDDING_MODEL_NAME

    # Convertir un seul texte en liste
    if isinstance(texts, str):
        texts = [texts]

    # Troncature de sécurité pour éviter les dépassements de limite
    safe_texts = []
    for text in texts:
        # Estimation simple: 1 token ≈ 4 chars
        token_count = len(text) // 4

        # Limite API: ~8191 tokens pour text-embedding-3-small
        if token_count > 8000:
            safe_text = text[:32000]  # ~8000 tokens
        else:
            safe_text = text
        safe_texts.append(safe_text)

    try:
        # Utilisation de la syntaxe client OpenAI native
        response = client.embeddings.create(
            model=model,
            input=safe_texts,
            **kwargs
        )
        logger.info(f"Embedding réussi pour {len(texts)} texte(s) avec le modèle {model}")

        # Extraire les embeddings de la réponse
        embeddings = [item.embedding for item in response.data]
        return embeddings

    except Exception as e:
        logger.error(f"Erreur lors de l'appel embedding: {e}")
        raise

# ==========================================
# EXTRACTION DE TEXTE (OCR)
# ==========================================

def extract_text_from_image(image_path: Union[str, Path], lang: str = None) -> str:
    """
    Extrait le texte d'une image avec Tesseract OCR (gratuit, sans API)

    Args:
        image_path: Chemin vers l'image
        lang: Langue(s) pour l'OCR (défaut: config.OCR_LANGUAGE)
              Format: 'fra' pour français, 'eng' pour anglais, 'fra+eng' pour les deux

    Returns:
        Texte extrait de l'image

    Raises:
        Exception: Si l'OCR échoue
    """
    if not config.ENABLE_IMAGE_OCR:
        logger.info("OCR désactivé")
        return "[Image présente - OCR désactivé]"

    lang = lang or config.OCR_LANGUAGE

    try:
        # Ouvrir l'image
        img = Image.open(image_path)

        # Amélioration de la qualité pour OCR
        # Upscaling si l'image est trop petite
        min_dim = config.MIN_OCR_DIMENSION
        if min(img.size) < min_dim:
            ratio = min_dim / min(img.size)
            new_size = tuple(int(dim * ratio) for dim in img.size)
            img = img.resize(new_size, Image.Resampling.LANCZOS)
            logger.info(f"Image agrandie à {new_size} pour améliorer l'OCR")

        # Conversion en niveaux de gris pour améliorer la reconnaissance
        if img.mode != 'L':
            img = img.convert('L')

        # Extraction du texte avec Tesseract
        text = pytesseract.image_to_string(img, lang=lang)

        # Nettoyer le texte extrait
        text = text.strip()

        if text:
            logger.info(f"OCR réussi : {len(text)} caractères extraits de {image_path}")
            return text
        else:
            logger.warning(f"Aucun texte détecté dans l'image : {image_path}")
            return "[Image sans texte détectable]"

    except pytesseract.TesseractNotFoundError:
        error_msg = (
            "Tesseract OCR n'est pas installé. "
            "Installation requise : sudo apt-get install tesseract-ocr tesseract-ocr-fra"
        )
        logger.error(error_msg)
        return f"[Erreur OCR : Tesseract non installé]"

    except Exception as e:
        logger.error(f"Erreur lors de l'OCR de l'image {image_path}: {e}")
        return f"[Erreur OCR : {str(e)}]"

# ==========================================
# EXTRACTION EXCEL
# ==========================================

def extract_text_from_excel(file_path: Union[str, Path], max_sheets: int = None, strategy: str = None) -> str:
    """
    Extrait le texte d'un fichier Excel (.xlsx, .xlsm) avec sélection intelligente des onglets.

    Args:
        file_path: Chemin vers le fichier Excel
        max_sheets: Nombre maximum d'onglets à extraire (défaut: config.MAX_EXCEL_SHEETS)
        strategy: Stratégie de sélection ('exact', 'auto' ou 'first')
                  - 'exact': Correspondance stricte du nom d'onglet (recommandé)
                  - 'auto': Recherche partielle dans le nom d'onglet
                  - 'first': Prend les N premiers onglets

    Returns:
        Texte formaté extrait des onglets sélectionnés

    Raises:
        Exception: Si la lecture échoue
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        error_msg = "openpyxl non installé. Installation requise : pip install openpyxl"
        logger.error(error_msg)
        return f"[Erreur Excel : openpyxl non installé]"

    max_sheets = max_sheets or config.MAX_EXCEL_SHEETS
    strategy = strategy or config.EXCEL_SHEET_STRATEGY
    max_rows = config.MAX_EXCEL_ROWS

    try:
        # Charger le workbook (data_only=True pour récupérer les valeurs calculées)
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        sheet_names = workbook.sheetnames

        logger.info(f"Fichier Excel chargé : {len(sheet_names)} onglet(s) trouvé(s)")

        if not sheet_names:
            return "[Alerte : Aucun onglet trouvé dans le fichier Excel]"

        # Sélection des onglets selon la stratégie
        selected_sheets = []

        if strategy == "exact":
            # Stratégie exacte : correspondance stricte du nom d'onglet
            priority_names = [name.strip() for name in config.EXCEL_PRIORITY_SHEETS.split(",")]

            # Chercher les onglets avec correspondance exacte
            for priority in priority_names:
                for sheet_name in sheet_names:
                    # Correspondance exacte (insensible à la casse)
                    if sheet_name.strip().lower() == priority.lower() and sheet_name not in selected_sheets:
                        selected_sheets.append(sheet_name)
                        logger.info(f"Onglet trouvé (correspondance exacte) : {sheet_name}")
                        if len(selected_sheets) >= max_sheets:
                            break
                if len(selected_sheets) >= max_sheets:
                    break

            # Si on n'a pas trouvé tous les onglets, loguer un avertissement
            if len(selected_sheets) < len(priority_names):
                missing = [p for p in priority_names if p not in [s.strip() for s in selected_sheets]]
                logger.warning(f"Onglets non trouvés : {', '.join(missing)}")

        elif strategy == "auto":
            # Stratégie intelligente : recherche partielle dans le nom
            priority_names = [name.strip() for name in config.EXCEL_PRIORITY_SHEETS.split(",")]

            # 1. Chercher les onglets prioritaires (recherche partielle)
            for priority in priority_names:
                for sheet_name in sheet_names:
                    if priority.lower() in sheet_name.lower() and sheet_name not in selected_sheets:
                        selected_sheets.append(sheet_name)
                        if len(selected_sheets) >= max_sheets:
                            break
                if len(selected_sheets) >= max_sheets:
                    break

            # 2. Compléter avec les premiers onglets si besoin
            if len(selected_sheets) < max_sheets:
                for sheet_name in sheet_names:
                    if sheet_name not in selected_sheets:
                        selected_sheets.append(sheet_name)
                        if len(selected_sheets) >= max_sheets:
                            break
        else:
            # Stratégie 'first' : prendre les N premiers onglets
            selected_sheets = sheet_names[:max_sheets]

        logger.info(f"Onglets sélectionnés : {selected_sheets}")

        # Extraction du contenu
        content = f"[FICHIER EXCEL: {Path(file_path).name}]\n"
        content += f"Nombre total d'onglets : {len(sheet_names)}\n"
        content += f"Onglets analysés : {', '.join(selected_sheets)}\n\n"

        for sheet_name in selected_sheets:
            try:
                sheet = workbook[sheet_name]
                content += f"\n{'='*60}\n"
                content += f"ONGLET : {sheet_name}\n"
                content += f"{'='*60}\n\n"

                # Compter les lignes avec données
                rows_with_data = 0
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows_with_data += 1

                # Extraire les données (limité par MAX_EXCEL_ROWS)
                row_count = 0
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    # Limiter le nombre de lignes
                    if max_rows > 0 and row_count >= max_rows:
                        content += f"\n... (lignes suivantes ignorées, limite : {max_rows} lignes)\n"
                        break

                    # Ignorer les lignes complètement vides
                    if not any(cell is not None for cell in row):
                        continue

                    row_count += 1

                    # Formater la ligne
                    row_text = []
                    for col_idx, cell_value in enumerate(row, start=1):
                        if cell_value is not None:
                            # Convertir en string et nettoyer
                            cell_str = str(cell_value).strip()
                            if cell_str:
                                col_letter = get_column_letter(col_idx)
                                row_text.append(f"{col_letter}: {cell_str}")

                    if row_text:
                        content += f"Ligne {row_idx}: " + " | ".join(row_text) + "\n"

                content += f"\nTotal lignes avec données : {rows_with_data}\n"

            except Exception as sheet_err:
                logger.warning(f"Erreur lors de la lecture de l'onglet '{sheet_name}': {sheet_err}")
                content += f"\n[Erreur lors de la lecture de l'onglet '{sheet_name}']\n"

        workbook.close()
        logger.info(f"Extraction Excel terminée : {len(selected_sheets)} onglet(s) traité(s)")

        return content

    except Exception as e:
        logger.error(f"Erreur lors de la lecture du fichier Excel {file_path}: {e}")
        return f"[Erreur lecture Excel : {str(e)}]"

# ==========================================
# GESTION DES TOKENS
# ==========================================

def estimate_tokens(text: str, model: str = None) -> int:
    """
    Estimation du nombre de tokens (approximation simple)

    Args:
        text: Texte à analyser
        model: Nom du modèle (optionnel, non utilisé pour l'approximation)

    Returns:
        Nombre de tokens (estimation)

    Note:
        Utilise une approximation simple (1 token ≈ 4 caractères) pour éviter les
        problèmes avec token_counter qui ne supporte pas les proxies custom
    """
    # Approximation simple et fiable : 1 token ≈ 4 caractères
    return max(len(text) // 4, 1)

# ==========================================
# CALCUL DES COÛTS
# ==========================================

def calculate_cost(input_tokens: int, output_tokens: int = 0, model: str = None) -> dict:
    """
    Calcule le coût estimé pour un appel API

    Args:
        input_tokens: Nombre de tokens en entrée
        output_tokens: Nombre de tokens en sortie (0 par défaut)
        model: Nom du modèle (optionnel, utilise MODEL_NAME par défaut)

    Returns:
        Dictionnaire avec coûts détaillés:
        {
            "input_cost": float,      # Coût entrée en USD
            "output_cost": float,     # Coût sortie en USD
            "total_cost": float,      # Coût total en USD
            "input_tokens": int,      # Tokens entrée
            "output_tokens": int,     # Tokens sortie
            "total_tokens": int,      # Total tokens
            "model": str              # Modèle utilisé
        }
    """
    model = model or config.MODEL_NAME

    # Récupérer les tarifs du modèle
    pricing = config.PRICING.get(model, config.DEFAULT_PRICING)

    # Calcul des coûts (tarifs par 1000 tokens)
    input_cost = (input_tokens / 1000) * pricing["input"]
    output_cost = (output_tokens / 1000) * pricing["output"]
    total_cost = input_cost + output_cost

    return {
        "input_cost": input_cost,
        "output_cost": output_cost,
        "total_cost": total_cost,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "total_tokens": input_tokens + output_tokens,
        "model": model
    }

def format_cost(cost: float) -> str:
    """
    Formate un coût en USD pour l'affichage

    Args:
        cost: Coût en USD

    Returns:
        Chaîne formatée (ex: "$0.0015" ou "$1.50")
    """
    if cost < 0.01:
        return f"${cost:.6f}"
    elif cost < 1.0:
        return f"${cost:.4f}"
    else:
        return f"${cost:.2f}"

# ==========================================
# CHARGEMENT DES PROMPTS
# ==========================================

def load_prompt(filename: str) -> str:
    """
    Charge un prompt depuis le dossier prompts/

    Args:
        filename: Nom du fichier prompt (ex: 'app_system_prompt.txt')

    Returns:
        Contenu du prompt

    Raises:
        FileNotFoundError: Si le fichier n'existe pas
    """
    prompt_path = Path(__file__).parent / "prompts" / filename

    try:
        with open(prompt_path, "r", encoding="utf-8") as f:
            prompt = f.read().strip()
        logger.info(f"Prompt chargé : {filename}")
        return prompt
    except Exception as e:
        logger.error(f"Impossible de charger le prompt {filename}: {e}")
        raise
